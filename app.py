import asyncio
import io
import json
import logging
import re
import os # Import OS to read env vars
import pandas as pd
from aiogram import Bot, Dispatcher, F, types
from aiogram.filters import CommandStart
from aiogram.types import BufferedInputFile
from datetime import datetime, timezone, timedelta
# Imports for Styling
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Imports for Environment Variables
from dotenv import load_dotenv

# Imports for Keep-Alive Web Server
from aiohttp import web

# ------------------------------------------------------------------------------
# CONFIGURATION
# ------------------------------------------------------------------------------

# 1. Load environment variables from .env file (for local development)
load_dotenv()

# 2. Get Token from Environment
BOT_TOKEN = os.getenv("BOT_TOKEN")

if not BOT_TOKEN:
    raise ValueError("No BOT_TOKEN found in environment variables!")

# Initialize Bot and Dispatcher
bot = Bot(token=BOT_TOKEN)
dp = Dispatcher()

# Configure logging
logging.basicConfig(level=logging.INFO)

# Global variable to store the bot's start time
BOT_START_TIME = datetime.now(timezone.utc)

# ------------------------------------------------------------------------------
# KEEP-ALIVE WEB SERVER
# ------------------------------------------------------------------------------

async def health_check(request):
    from datetime import datetime, timezone, timedelta

    dhaka_tz = timezone(timedelta(hours=6))
    now_dhaka = datetime.now(dhaka_tz).strftime("%Y-%m-%d %H:%M:%S")

    start_time = request.app.get("start_time")
    if start_time:
        delta = datetime.now(timezone.utc) - start_time
        days = delta.days
        hours, r = divmod(delta.seconds, 3600)
        mins, secs = divmod(r, 60)
        uptime = f"{days}d {hours}h {mins}m {secs}s"
    else:
        uptime = "Not Available"

    html = f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>Bot Health — Online</title>
<meta name="viewport" content="width=device-width, initial-scale=1">

<style>
    body {{
        margin:0;
        font-family: Inter, sans-serif;
        background:#0d0f18;
        color:white;
        height:100vh;
        display:flex;
        justify-content:center;
        align-items:center;
        padding:20px;
    }}

    .card {{
        width:100%;
        max-width:500px;
        background:rgba(255,255,255,0.05);
        border:1px solid rgba(255,255,255,0.10);
        border-radius:16px;
        padding:25px 20px;
        box-sizing:border-box;
    }}

    .title {{
        text-align:center;
        font-size:24px;
        font-weight:700;
        margin-bottom:6px;
    }}

    .subtitle {{
        text-align:center;
        font-size:13px;
        color:#cccccc;
        margin-bottom:25px;
    }}

    .status {{
        display:flex;
        align-items:center;
        justify-content:center;
        gap:10px;
        margin-bottom:18px;
        font-size:16px;
    }}

    .dot {{
        width:14px;
        height:14px;
        border-radius:50%;
        background:#00e676;
        box-shadow:0 0 6px #00e676;
    }}

    .info {{
        display:flex;
        flex-direction:column;
        gap:10px;
    }}

    .item {{
        padding:12px;
        border-radius:10px;
        background:rgba(255,255,255,0.04);
        border:1px solid rgba(255,255,255,0.06);
        font-size:14px;
    }}

    .label {{
        font-weight:600;
        margin-right:4px;
    }}

</style>
</head>

<body>

<div class="card">
    <div class="title">🤖 Bot Status</div>
    <div class="subtitle">System is running normally.</div>

    <div class="status">
        <div class="dot"></div>
        <span>ONLINE</span>
    </div>

    <div class="info">
        <div class="item"><span class="label">Uptime:</span> {uptime}</div>
        <div class="item"><span class="label">Current Time:</span> {now_dhaka}</div>
        <div class="item"><span class="label">Version:</span> v1.0.0</div>
    </div>
</div>

</body>
</html>"""

    return web.Response(text=html, content_type="text/html")


async def start_web_server(start_time: datetime):
    """
    Starts a small aiohttp web server on the port defined by Render.
    The start_time is passed to the app context for uptime calculation.
    """
    app = web.Application()
    # FIX: Set the start_time on the app context
    app["start_time"] = start_time
    
    app.router.add_get('/', health_check)
    
    # Render provides the PORT environment variable. Default to 8080 if not found.
    port = int(os.getenv("PORT", 10000))
    
    runner = web.AppRunner(app)
    await runner.setup()
    # 0.0.0.0 allows external access (required for Render)
    site = web.TCPSite(runner, "0.0.0.0", port)
    await site.start()
    logging.info(f"Web server started on port {port}")

# ------------------------------------------------------------------------------
# STYLING FUNCTION (Unchanged)
# ------------------------------------------------------------------------------

def create_styled_excel(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        data_font = Font(name='Segoe UI', size=10)
        data_alignment = Alignment(horizontal='center', vertical='center')
        thin_border_side = Side(border_style='thin', color='000000')
        thin_border = Border(left=thin_border_side, right=thin_border_side, 
                             top=thin_border_side, bottom=thin_border_side)

        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for i, column in enumerate(worksheet.columns):
            max_length = 0
            column_letter = get_column_letter(i + 1)
            header_val = column[0].value
            if header_val: max_length = len(str(header_val))

            for cell in column:
                if cell.row > 1:
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = thin_border
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except: pass
            
            adjusted_width = (max_length + 4) * 1.1
            if adjusted_width > 50: adjusted_width = 50
            worksheet.column_dimensions[column_letter].width = adjusted_width

    output.seek(0)
    return output

# ------------------------------------------------------------------------------
# LOGIC FUNCTIONS (Unchanged)
# ------------------------------------------------------------------------------

def get_timestamp():
    # Add +6 hours to current time
    time_plus_6 = datetime.now() + timedelta(hours=6)
    return time_plus_6.strftime("%d_%m_%Y_%I_%M_%S_%p")


def is_phone_number(s):
    if not s: return False
    s = str(s).strip()
    return bool(re.match(r'^\+?\d+$', s))

def process_json_data(json_content):
    try:
        data = json.loads(json_content)
    except json.JSONDecodeError:
        return None, None

    if not isinstance(data, list):
        return None, None

    number_rows = []
    email_rows = []

    for entry in data:
        email_field = str(entry.get('email', '')).strip()
        username = entry.get('username', '')
        password = entry.get('password', '')
        auth_code = entry.get('auth_code', '')

        if is_phone_number(email_field):
            number_rows.append({'Username': username, 'Password': password, '2FA': auth_code, 'Number': email_field})
        else:
            email_rows.append({'Username': username, 'Password': password, '2FA': auth_code, 'Email': email_field})

    buffer_num = None
    buffer_mail = None

    if number_rows:
        df_num = pd.DataFrame(number_rows)[['Username', 'Password', '2FA', 'Number']]
        buffer_num = create_styled_excel(df_num)

    if email_rows:
        df_mail = pd.DataFrame(email_rows)[['Username', 'Password', '2FA', 'Email']]
        buffer_mail = create_styled_excel(df_mail)

    return buffer_num, buffer_mail

# ------------------------------------------------------------------------------
# HANDLERS (Unchanged)
# ------------------------------------------------------------------------------

@dp.message(CommandStart())
async def cmd_start(message: types.Message):
    await message.answer(
        "👋 <b>Instagram Data Processor</b>\n\n"
        "Send me a <b>.json</b> file and I will convert it into formatted Excel sheets.\n\n"
        "🔹 <b>Phone List:</b> Only numbers in email field.\n"
        "🔹 <b>Mail List:</b> Emails or empty fields.\n\n"
        "<i>Reply to any file to re-process it!</i>",
        parse_mode="HTML"
    )

async def handle_document_processing(message: types.Message, document: types.Document):
    status_msg = await message.answer("🎨 <b>Styling and processing Excel files...</b>", parse_mode="HTML")
    try:
        file_io = io.BytesIO()
        await bot.download(document, destination=file_io)
        file_content = file_io.read().decode('utf-8')
        num_excel, mail_excel = process_json_data(file_content)

        if num_excel is None and mail_excel is None:
            await status_msg.edit_text("❌ Error: Invalid JSON or empty data.")
            return

        timestamp = get_timestamp()
        files_to_send = []
        if num_excel:
            files_to_send.append(BufferedInputFile(num_excel.read(), filename=f"number_instagram_{timestamp}.xlsx"))
        if mail_excel:
            files_to_send.append(BufferedInputFile(mail_excel.read(), filename=f"mail_instagram_{timestamp}.xlsx"))

        if not files_to_send:
            await status_msg.edit_text("⚠️ No valid data found.")
        else:
            await status_msg.delete()
            for file_obj in files_to_send:
                await message.reply_document(document=file_obj, caption="✅ <b>Here is your file</b>", parse_mode="HTML")

    except Exception as e:
        logging.error(f"Error: {e}")
        await status_msg.edit_text(f"❌ Error: {str(e)}")

@dp.message(F.document)
async def on_document(message: types.Message):
    if message.document.file_name and message.document.file_name.endswith('.json'):
        await handle_document_processing(message, message.document)
    else:
        await message.reply("⚠️ Please send a <b>.json</b> file.", parse_mode="HTML")

@dp.message(F.reply_to_message & F.reply_to_message.document)
async def on_reply_to_document(message: types.Message):
    original_doc = message.reply_to_message.document
    if original_doc.file_name and original_doc.file_name.endswith('.json'):
        await handle_document_processing(message, original_doc)
    else:
        await message.reply("⚠️ The original message is not a .json file.")

# ------------------------------------------------------------------------------
# MAIN
# ------------------------------------------------------------------------------
async def main():
    # Start the keep-alive web server first, passing the bot's start time
    await start_web_server(BOT_START_TIME)
    
    print("Bot is running with styled Excel output...")
    # Start the bot polling
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())